"""Create XLSX spreadsheet files for download.

Tier: std-create (Tier 1) — single-shot file generation.

**std-create**: build an Excel from `columns + rows` in one call. Used
for KPI exports, table dumps, anything where the agent already has the
data structured. Auto-loaded for every `agentino.tools.std` agent.

For multi-step workbook work — open existing files, fill templates,
edit cells, multi-sheet formulas — the convention is **Tier 2 /
`skill-xlsx`**: a per-tenant skill bundle, opt-in via `skills:` in that
tenant's agents.yml. Tier 2 isn't centralized — each tenant ships only
what its agents need.

Tool name in agents.yml: `create_spreadsheet` (back-compat).
"""

import io
from datetime import datetime

from agentino.core.tool import tool

from .storage import get_default_store


@tool
async def create_spreadsheet(
    title: str,
    columns: list[str],
    rows: list[list[str]],
    filename: str = "",
    sheet_name: str = "Sheet1",
) -> str:
    """Create an Excel (.xlsx) spreadsheet and attach it for download. Use when the user asks to export data as Excel, spreadsheet, or XLSX.

    Args:
        title: Document title (used in filename if filename not provided)
        columns: Column headers
        rows: Data rows (list of lists, each inner list is one row)
        filename: Optional output filename (without extension). Defaults to title-based name.
        sheet_name: Worksheet name (default: Sheet1)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    header_font = Font(bold=True, size=11, color="374151")
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    data_font = Font(size=11, color="1F2937")
    stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    # Header row
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Detect numeric columns from first data row
    numeric_cols = set()
    if rows:
        import re

        for ci, val in enumerate(rows[0]):
            if re.match(r"^[+-]?[€$£]?[\d,]+\.?\d*%?$", str(val).strip()):
                numeric_cols.add(ci)

    # Data rows
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = data_font
            cell.border = thin_border
            if (ri - 2) % 2 == 1:
                cell.fill = stripe_fill
            if (ci - 1) in numeric_cols:
                cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    for ci in range(1, len(columns) + 1):
        max_len = len(str(columns[ci - 1]))
        for ri in range(2, min(len(rows) + 2, 102)):
            cell_val = ws.cell(row=ri, column=ci).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)

    # Freeze header row + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    if not filename:
        safe = title.lower().replace(" ", "-").replace("/", "-")[:40]
        filename = safe
    now = datetime.now().strftime("%Y-%m-%d")
    final_filename = f"{filename}-{now}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    content_bytes = buf.getvalue()

    # Indexable preview: header row + up to 100 data rows as TSV.
    preview_rows = ["\t".join(columns)]
    for r in rows[:100]:
        preview_rows.append("\t".join(str(c) for c in r))
    extracted = "\n".join(preview_rows)[:50_000]

    try:
        stored = get_default_store().save(
            content_bytes=content_bytes,
            filename=final_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            title=title,
            extracted_text=extracted,
        )
    except Exception as e:
        return f"Spreadsheet generation failed: {e}"

    row_count = len(rows)
    return (
        f"Spreadsheet created ({row_count} rows, {len(columns)} columns). "
        f"Download: [{stored.filename}]({stored.url})"
    )
